import os
from typing import List, Tuple

from openpyxl import Workbook


def get_training_data() -> Tuple[List[List[int]], List[int]]:
    """Return sample training data matching build_model.py arrays.

    Features: [Tajwid, Kelancaran, Kefasihan] in 1-100 scale
    Target: NilaiAkhir (1-100)
    """
    X_train = [
        [95, 95, 95],
        [90, 95, 90],
        [95, 90, 95],
        [85, 90, 85],
        [80, 85, 80],
        [85, 80, 85],
        [75, 80, 75],
        [75, 80, 75],
        [70, 75, 70],
        [75, 70, 75],
        [65, 70, 65],
        [60, 65, 60],
        [70, 60, 70],
        [55, 60, 55],
        [50, 55, 50],
        [45, 50, 45],
        [40, 45, 40],
        [35, 40, 35],
        [30, 35, 30],
        [25, 30, 25],
        [20, 25, 20],
        [15, 20, 15],
        [10, 15, 10],
        [5, 10, 5],
        [1, 5, 1],
    ]
    y_train = [
        95, 92, 93,
        87, 82, 83,
        77, 75,
        72, 73, 67,
        62, 67,
        57, 52, 47,
        42, 37, 32,
        27, 22, 17,
        12, 7, 2,
    ]
    return X_train, y_train


def build_excel_template(output_path: str) -> None:
    X_train, y_train = get_training_data()

    wb = Workbook()

    # Sheet: Data
    ws_data = wb.active
    ws_data.title = "Data"
    ws_data["A1"] = "Tajwid"
    ws_data["B1"] = "Kelancaran"
    ws_data["C1"] = "Kefasihan"
    ws_data["D1"] = "NilaiAkhir"
    ws_data["E1"] = "Kelas"
    ws_data["F1"] = "Tajwid2"
    ws_data["G1"] = "Kelancaran2"
    ws_data["H1"] = "Kefasihan2"

    for idx, (features, target) in enumerate(zip(X_train, y_train), start=2):
        tajwid, kelancaran, kefasihan = features
        ws_data.cell(row=idx, column=1, value=tajwid)
        ws_data.cell(row=idx, column=2, value=kelancaran)
        ws_data.cell(row=idx, column=3, value=kefasihan)
        ws_data.cell(row=idx, column=4, value=target)
        ws_data.cell(row=idx, column=5, value=f"=IF(D{idx}>=75,\"Lulus\",\"Tidak Lulus\")")
        ws_data.cell(row=idx, column=6, value=f"=A{idx}^2")
        ws_data.cell(row=idx, column=7, value=f"=B{idx}^2")
        ws_data.cell(row=idx, column=8, value=f"=C{idx}^2")

    last_row = 1 + len(X_train)

    # Sheet: Parameter
    ws_param = wb.create_sheet("Parameter")
    headers = [
        "Kelas",
        "Count",
        "Prior",
        "Tajwid_mu",
        "Tajwid_var",
        "Kelancaran_mu",
        "Kelancaran_var",
        "Kefasihan_mu",
        "Kefasihan_var",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws_param.cell(row=1, column=col_idx, value=header)

    ws_param["A2"] = "Lulus"
    ws_param["A3"] = "Tidak Lulus"

    # Ranges in Data sheet
    rng_A = f"Data!$A$2:$A${last_row}"
    rng_B = f"Data!$B$2:$B${last_row}"
    rng_C = f"Data!$C$2:$C${last_row}"
    rng_D = f"Data!$D$2:$D${last_row}"
    rng_E = f"Data!$E$2:$E${last_row}"
    rng_F = f"Data!$F$2:$F${last_row}"
    rng_G = f"Data!$G$2:$G${last_row}"
    rng_H = f"Data!$H$2:$H${last_row}"

    # Counts
    ws_param["B2"] = f"=COUNTIF({rng_E},$A2)"
    ws_param["B3"] = f"=COUNTIF({rng_E},$A3)"

    # Priors
    ws_param["C2"] = "=B2/(B2+B3)"
    ws_param["C3"] = "=B3/(B2+B3)"

    # Means and variances for each feature by class
    # Tajwid
    ws_param["D2"] = f"=SUMIFS({rng_A},{rng_E},$A2)/$B2"
    ws_param["E2"] = f"=SUMIFS({rng_F},{rng_E},$A2)/$B2 - D2^2"
    ws_param["D3"] = f"=SUMIFS({rng_A},{rng_E},$A3)/$B3"
    ws_param["E3"] = f"=SUMIFS({rng_F},{rng_E},$A3)/$B3 - D3^2"

    # Kelancaran
    ws_param["F2"] = f"=SUMIFS({rng_B},{rng_E},$A2)/$B2"
    ws_param["G2"] = f"=SUMIFS({rng_G},{rng_E},$A2)/$B2 - F2^2"
    ws_param["F3"] = f"=SUMIFS({rng_B},{rng_E},$A3)/$B3"
    ws_param["G3"] = f"=SUMIFS({rng_G},{rng_E},$A3)/$B3 - F3^2"

    # Kefasihan
    ws_param["H2"] = f"=SUMIFS({rng_C},{rng_E},$A2)/$B2"
    ws_param["I2"] = f"=SUMIFS({rng_H},{rng_E},$A2)/$B2 - H2^2"
    ws_param["H3"] = f"=SUMIFS({rng_C},{rng_E},$A3)/$B3"
    ws_param["I3"] = f"=SUMIFS({rng_H},{rng_E},$A3)/$B3 - H3^2"

    # Sheet: Prediksi
    ws_pred = wb.create_sheet("Prediksi")
    ws_pred["A1"] = "Input"
    ws_pred["A2"] = "Tajwid"
    ws_pred["A3"] = "Kelancaran"
    ws_pred["A4"] = "Kefasihan"

    ws_pred["A6"] = "LogP_Lulus"
    ws_pred["A7"] = "LogP_Tidak"
    ws_pred["A9"] = "Prediksi"
    ws_pred["A10"] = "Prob_Lulus"

    # Log posterior for Lulus (row 2 in Parameter)
    ws_pred["B6"] = (
        "=LN(Parameter!$C$2)"
        " + (-0.5*LN(2*PI()*MAX(Parameter!$E$2,1E-6)) - (B2-Parameter!$D$2)^2/(2*MAX(Parameter!$E$2,1E-6)))"
        " + (-0.5*LN(2*PI()*MAX(Parameter!$G$2,1E-6)) - (B3-Parameter!$F$2)^2/(2*MAX(Parameter!$G$2,1E-6)))"
        " + (-0.5*LN(2*PI()*MAX(Parameter!$I$2,1E-6)) - (B4-Parameter!$H$2)^2/(2*MAX(Parameter!$I$2,1E-6)))"
    )

    # Log posterior for Tidak Lulus (row 3 in Parameter)
    ws_pred["B7"] = (
        "=LN(Parameter!$C$3)"
        " + (-0.5*LN(2*PI()*MAX(Parameter!$E$3,1E-6)) - (B2-Parameter!$D$3)^2/(2*MAX(Parameter!$E$3,1E-6)))"
        " + (-0.5*LN(2*PI()*MAX(Parameter!$G$3,1E-6)) - (B3-Parameter!$F$3)^2/(2*MAX(Parameter!$G$3,1E-6)))"
        " + (-0.5*LN(2*PI()*MAX(Parameter!$I$3,1E-6)) - (B4-Parameter!$H$3)^2/(2*MAX(Parameter!$I$3,1E-6)))"
    )

    # Decision and probability
    ws_pred["B9"] = "=IF(B6>B7,\"Lulus\",\"Tidak Lulus\")"
    ws_pred["B10"] = "=1/(1+EXP(B7-B6))"

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    output_file = os.path.join("output", "Template_GNB.xlsx")
    build_excel_template(output_file)
    print(f"Excel template created at: {output_file}")







